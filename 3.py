import streamlit as st
import pdfplumber
import re
import csv
from openpyxl import Workbook

def extract_data_from_pdf(file_path):
    awb_regex = r'\b\d{14,15}\b'
    order_id_regex = r'Order No: (\d+_[^,\s]+)'
    date_regex = r'Date : (\d{2}-\d{2}-\d{4})'
    time_regex = r'Time : (\d{2}:\d{2}:\d{2})'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet['A1'] = 'AWB Number'
    worksheet['B1'] = 'Order Id'
    worksheet['C1'] = 'Date'
    worksheet['D1'] = 'Time'

    with pdfplumber.open(file_path) as pdf:
        pdf_text = ''
        awb_numbers = []
        order_ids = []
        dates = []
        times = []

        for page in pdf.pages:
            pdf_text += page.extract_text()

            awb_numbers = re.findall(awb_regex, pdf_text)
            order_ids = re.findall(order_id_regex, pdf_text)
            dates = re.findall(date_regex, pdf_text)
            times = re.findall(time_regex, pdf_text)

        for row, (awb, order_id) in enumerate(zip(awb_numbers, order_ids), start=2):
            worksheet.cell(row=row, column=1, value=awb)
            worksheet.cell(row=row, column=2, value=order_id)

        for row in range(2, worksheet.max_row+1):
            worksheet.cell(row=row, column=3, value=dates[0])
            worksheet.cell(row=row, column=4, value=times[0])

        return workbook

def main():
    st.title('PDF Data Extractor')

    uploaded_file = st.file_uploader('Upload a PDF file', type=['pdf'])

    if uploaded_file is not None:
        workbook = extract_data_from_pdf(uploaded_file)

        output_file = f'output.xlsx'
        workbook.save(output_file)

        st.download_button(
            label='Download Excel file',
            data=workbook_to_bytes(workbook),
            file_name=output_file,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

def workbook_to_bytes(workbook):
    with open('temp.xlsx', 'wb') as f:
        workbook.save(f)

    with open('temp.xlsx', 'rb') as f:
        data = f.read()

    return data

if __name__ == '__main__':
    main()
